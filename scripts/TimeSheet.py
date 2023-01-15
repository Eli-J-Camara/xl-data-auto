from openpyxl import load_workbook
import datetime
from populate import get_punch_times

def ws_header():
    # I've changed these dates to match the date range of the csv file. Also make sure that the date format does not
    # change on the csv file.
    weekly_timesheet = 'apr-weekly-timesheet-dec-5th.xlsx'
    source_data_ws = load_workbook(weekly_timesheet)['Sheet']
    date_range = source_data_ws['B2'].value
    from_date = date_range[0:5]
    to_date = date_range[14:19]
    log_heading = f'Time Reporting Summary ({from_date} - {to_date}):'

    # Header date numberings
    from_range = int(from_date[3:5])
    to_range = int(to_date[3:5])
    header_dates = [date for date in range(from_range, to_range + 1)]
    header_date_column = ['A', 'E', 'I', 'M', 'Q']
    header_date_row = ['8', '19', '27', '35', '43']
    header_cells = [[col + row for row in header_date_row] for col in header_date_column]
    return log_heading, header_cells, header_dates

class TimeSheet:

    def __init__(self):
        self.months = {1: 'January', 2: 'February', 3: 'March', 4: 'April', 5: 'May', 6: 'June', 7: 'July', 8: 'August',
                       9: 'September', 10: 'October', 11: 'November', 12: 'December'}
        self.title = self.create_ws_title()
        self.header = ws_header()
        self.excel_template = 'TimeReportLog(Oct2022).xlsx'

    def create_ws_title(self):
        # """Generates dynamic work sheet title."""
        weekly_timesheet = 'apr-weekly-timesheet-dec-5th.xlsx'
        source_data_ws = load_workbook(weekly_timesheet)['Sheet']
        date_range = source_data_ws['B2'].value
        day = date_range[3:5]
        month_num = date_range[0:2]
        current_month = self.months[int(month_num)]
        ws_name = f'Weekly Time Reporting {current_month[0:3]}. {day}'
        return ws_name

    def generate_work_sheet(self):
        wb = load_workbook(filename=self.excel_template)
        template_ws = wb['Weekly Time Reporting Template']
        new_ws = wb.copy_worksheet(template_ws)
        new_ws.title = self.title
        new_ws['I1'] = self.header[0]
        for cells in self.header[1]:
            for i in range(5):
                index = self.header[1].index(cells)
                new_ws[cells[i]] = self.header[2][index]
        wb.save(self.excel_template)

    def populate_data(self):
        clock_in = get_punch_times('C', 'am')
        clock_out = get_punch_times('E', 'pm')

        wb = load_workbook(filename=self.excel_template)
        new_ws = wb[self.title]
        date_TimeColumn = {'A': 'C', 'E': 'G', 'I': 'K', 'M': 'O', 'Q': 'S'}
        for cell in new_ws['8']:
            new_row = 9
            value = cell.value
            if not bool(value):
                continue
            elif type(value) == int:
                coordinate = cell.coordinate
                for info in clock_in:
                    date = int(info['date'][-7:-5])
                    if value == date:
                        name_cell = coordinate[0] + str(new_row)
                        new_ws[name_cell] = info['name']
                        time_cell = date_TimeColumn[coordinate[0]] + str(new_row)
                        new_ws[time_cell] = info['clock-out'][0:-2]
                        new_row += 1
        wb.save(self.excel_template)