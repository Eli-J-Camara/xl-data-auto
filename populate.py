from openpyxl import load_workbook
from find_workbook import find_ws
import re


def get_punch_times(column, am_pm) -> list:
    """Returns a list of dicts with early clock in or late clock out names, dates and times. Example uses:
    get_punch_times('E', 'pm'), get_punch_times('C', 'am')"""
    weekly_timesheet = find_ws()
    if not weekly_timesheet:
        return []
    ws = load_workbook(weekly_timesheet)['Sheet']
    punch_times = []
    for cell in ws[column]:
        value = cell.value
        if not bool(value):
            continue
        if am_pm in value:
            left_of_colon = int(value[0:value.index(':')])
            if am_pm == 'pm':
                condition = left_of_colon > 5 and left_of_colon != 12
            else:
                condition = left_of_colon < 8
            if condition:
                punch_times.append({
                    'name': ws[f'A{cell.row}'].value,
                    'date': ws[f'B{cell.row}'].value,
                    'clock-out': value
                })
    return punch_times


# def get_missing_punches(col1, col2) -> list:
#     """Returns a list of dicts with the data needed for missing clock outs or clock ins. Example uses:
#     get_missing_punches('C', 'E') for missing clock outs and get_missing_punches('E', 'C') for missing clock ins"""
#     weekly_timesheet = find_ws()
#     if not weekly_timesheet:
#         return []
#     ws = load_workbook(weekly_timesheet)['Sheet']
#     punch_times = []
#     for cell in ws[col1]:
#         # regex to check that there's a time value in this cell
#         value = cell.value
#         if cell.value and bool(re.match(r'.*\d:\d{2}.*', value)):
#             if not ws[f'{col2}{cell.row}'].value:
#                 if col1 == 'C':
#                     punch_time = 'clock_in'
#                 else:
#                     punch_time = 'clock_out'
#                 punch_times.append({
#                     'name': ws[f'A{cell.row}'].value,
#                     'date': ws[f'B{cell.row}'].value,
#                     punch_time: value
#                 })
#     return punch_times
