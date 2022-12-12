# REMEMBER YOU ARE USING PYTHON3
from openpyxl import load_workbook, Workbook
from datetime import datetime


current_date = datetime.today().strftime('%Y-%m-%d')

# You will have to specify this to Zach's particular directory.

# new_file_path = f'/Users/lije/desktop/Software-Development/Chaz-Project/auto-ace/TimeReportingLog({current_date}).xlsx'
# wb = openpyxl.Workbook()
# wb.save(new_file_path)

# sheet_name = f"Weekly Time Report {current_date}"
# new_work_sheet = wb.create_sheet(sheet_name)
# template_log = openpyxl.load_workbook(template_file_path)
# source_ws = template_log['Weekly Time Reporting Oct. 17']

# Creating a copy of template.

excel_template = 'Time Reporting Log (October 2022).xlsx'

wb = load_workbook(filename=excel_template)

wb.save("NewReportingLog.xlsx")

#Find an alternative way to copy worksheet or excel file.

# maximum_row = source_ws.max_row
# maximum_column = source_ws.max_column

# for i in range(1, maximum_row + 1):
#     for j in range(1, maximum_column + 1):
#         #reading the cell value from template excel file
#         source_cell = source_ws.cell(row = i, column = j)
#         #writing the value read from the source file to the new file
#         new_cell = new_work_sheet.cell(row=i, column=j)
#         new_cell.value = source_cell.value



# template_log.save(template_file_path)


